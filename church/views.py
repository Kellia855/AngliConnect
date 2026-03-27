from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import ListView, CreateView, UpdateView
from django.urls import reverse_lazy, reverse
from django.http import JsonResponse, HttpResponse
from django.db.models import Count, Q
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth import update_session_auth_hash, login
from django.contrib.auth import logout as auth_logout
from django.contrib.auth.forms import PasswordChangeForm, AuthenticationForm
from django.contrib.auth.views import LoginView
from django.contrib import messages
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.views.decorators.cache import never_cache
from django.views.decorators.csrf import ensure_csrf_cookie
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle
from reportlab.lib.utils import ImageReader
import qrcode
from io import BytesIO
from datetime import timedelta
from .models import Parish, Member, Role, MemberRole, Diocese, Baptism, Confirmation, Marriage
from .forms import (
    MemberForm,
    RoleForm,
    MemberRoleForm,
    BaptismForm,
    ConfirmationForm,
    MarriageForm,
    ServiceSessionForm,
    AttendanceCheckInForm,
)
from .models import ServiceSession, AttendanceRecord


@method_decorator(never_cache, name='dispatch')
@method_decorator(ensure_csrf_cookie, name='dispatch')
class CustomLoginView(LoginView):
    """Custom login view that redirects members to portal and staff to dashboard"""
    template_name = 'registration/login.html'

    def get(self, request, *args, **kwargs):
        from django.middleware.csrf import get_token
        get_token(request)
        return super().get(request, *args, **kwargs)

    def get_success_url(self):
        if hasattr(self.request.user, 'member'):
            return reverse_lazy('church:member_portal')
        else:
            return reverse_lazy('church:dashboard')


def custom_logout(request):
    """Log out user and redirect to login page.

    This keeps frontend logout resilient even when a stale page token exists.
    """
    auth_logout(request)
    return redirect('login')


@login_required
def index(request):
    
    return render(request, 'church/index.html')


@login_required
def member_portal(request):

    if not hasattr(request.user, 'member'):
        return redirect('church:dashboard')
    
    member = request.user.member
    
  
    try:
        baptism = member.baptism
    except Baptism.DoesNotExist:
        baptism = None
    
    try:
        confirmation = member.confirmation
    except Confirmation.DoesNotExist:
        confirmation = None
    
    marriages = Marriage.objects.filter(
        Q(bride=member) | Q(groom=member)
    ).select_related('bride', 'groom', 'parish')
    
   
    member_roles = MemberRole.objects.filter(member=member).select_related('role')
    
    context = {
        'member': member,
        'baptism': baptism,
        'confirmation': confirmation,
        'marriages': marriages,
        'member_roles': member_roles,
    }
    
    return render(request, 'church/member_portal.html', context)


@login_required
def change_password(request):
    """Change password for authenticated users"""
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)  # Keep user logged in after password change
            messages.success(request, 'Your password was successfully updated!')
            
            if hasattr(request.user, 'member'):
                return redirect('church:member_portal')
            else:
                return redirect('church:dashboard')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = PasswordChangeForm(request.user)
    
    return render(request, 'church/change_password.html', {'form': form})



@login_required
def dashboard(request):
    """Analytics dashboard with church statistics"""
   
    total_members = Member.objects.count()
    total_dioceses = Diocese.objects.count()
    total_parishes = Parish.objects.count()
    total_roles = Role.objects.count()
    
    total_baptisms = Baptism.objects.count()
    total_confirmations = Confirmation.objects.count()
    total_marriages = Marriage.objects.count()

    today = timezone.localdate()
    today_sessions = ServiceSession.objects.filter(session_date=today)
    today_checkins = AttendanceRecord.objects.filter(session__session_date=today)
    today_checked_in = today_checkins.count()
    today_late = today_checkins.filter(status='LATE').count()
    
    
    diocese_stats = Diocese.objects.annotate(
        member_count=Count('parishes__members'),
        parish_count=Count('parishes')
    ).order_by('-member_count')
    
   
    parish_stats = Parish.objects.annotate(
        member_count=Count('members')
    ).select_related('diocese').order_by('-member_count')[:10]
    
    
    role_stats = Role.objects.annotate(
        assignment_count=Count('assignments', filter=Q(assignments__end_date__isnull=True))
    ).order_by('-assignment_count')
    
    
    recent_assignments = MemberRole.objects.select_related(
        'member', 'role'
    ).order_by('-start_date')[:10]
    

    recent_baptisms = Baptism.objects.select_related('member').order_by('-baptism_date')[:5]
    recent_confirmations = Confirmation.objects.select_related('member').order_by('-confirmation_date')[:5]
    recent_marriages = Marriage.objects.select_related('bride', 'groom').order_by('-marriage_date')[:5]
    
    context = {
        'total_members': total_members,
        'total_dioceses': total_dioceses,
        'total_parishes': total_parishes,
        'total_roles': total_roles,
        'total_baptisms': total_baptisms,
        'total_confirmations': total_confirmations,
        'total_marriages': total_marriages,
        'today_sessions_count': today_sessions.count(),
        'today_checked_in': today_checked_in,
        'today_late': today_late,
        'diocese_stats': diocese_stats,
        'parish_stats': parish_stats,
        'role_stats': role_stats,
        'recent_assignments': recent_assignments,
        'recent_baptisms': recent_baptisms,
        'recent_confirmations': recent_confirmations,
        'recent_marriages': recent_marriages,
    }
    
    return render(request, 'church/dashboard.html', context)


@login_required
def export_dashboard_excel(request):
    """Export attendance report """
    period = request.GET.get('period', 'monthly').lower()
    today = timezone.localdate()

    if period == 'weekly':
        period_label = 'Weekly'
        start_date = today - timedelta(days=6)
    elif period == 'annually':
        period_label = 'Annual'
        start_date = today.replace(month=1, day=1)
    else:
        period = 'monthly'
        period_label = 'Monthly'
        start_date = today.replace(day=1)

    end_date = today

    sessions_qs = ServiceSession.objects.filter(session_date__range=(start_date, end_date)).order_by('session_date')
    attendance_qs = AttendanceRecord.objects.filter(session__in=sessions_qs).select_related(
        'member',
        'session',
        'checked_in_by',
    )

    sessions_in_period = sessions_qs.count()
    checkins_in_period = attendance_qs.count()
    late_in_period = attendance_qs.filter(status='LATE').count()
    present_in_period = attendance_qs.filter(status='PRESENT').count()
    unique_members_in_period = attendance_qs.values('member_id').distinct().count()
    
    wb = Workbook()
    
    ws_overview = wb.active
    ws_overview.title = "Attendance Summary"
    
    
    header_fill = PatternFill(start_color="2C5F8D", end_color="2C5F8D", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    ws_overview['A1'] = "Attendance Report"
    ws_overview['A1'].font = Font(bold=True, size=14)
    ws_overview.merge_cells('A1:B1')

    ws_overview['A2'] = f"Report Period: {period_label} ({start_date} to {end_date})"
    ws_overview.merge_cells('A2:B2')
    
    ws_overview['A3'] = "Metric"
    ws_overview['B3'] = "Count"
    ws_overview['A3'].fill = header_fill
    ws_overview['B3'].fill = header_fill
    ws_overview['A3'].font = header_font
    ws_overview['B3'].font = header_font
    
    overview_data = [
        ["Attendance Sessions", sessions_in_period],
        ["Total Check-ins", checkins_in_period],
        ["Present", present_in_period],
        ["Late", late_in_period],
        ["Unique Members Checked In", unique_members_in_period],
    ]
    
    for idx, (metric, value) in enumerate(overview_data, start=4):
        ws_overview[f'A{idx}'] = metric
        ws_overview[f'B{idx}'] = value

    ws_period = wb.create_sheet("Session Breakdown")
    ws_period['A1'] = f"{period_label} Session Breakdown"
    ws_period['A1'].font = Font(bold=True, size=14)
    ws_period.merge_cells('A1:E1')
    ws_period['A2'] = f"Date Range: {start_date} to {end_date}"
    ws_period.merge_cells('A2:E2')

    session_headers = ['Date', 'Service Type', 'Status', 'Check-ins', 'Late']
    for col, title in zip(['A', 'B', 'C', 'D', 'E'], session_headers):
        ws_period[f'{col}4'] = title
        ws_period[f'{col}4'].fill = header_fill
        ws_period[f'{col}4'].font = header_font

    session_rows = []
    for session in sessions_qs:
        session_total = session.attendance_records.count()
        session_late = session.attendance_records.filter(status='LATE').count()
        session_rows.append([
            session.session_date,
            session.get_service_type_display(),
            session.get_status_display(),
            session_total,
            session_late,
        ])

    for idx, row in enumerate(session_rows, start=5):
        ws_period[f'A{idx}'] = row[0]
        ws_period[f'B{idx}'] = row[1]
        ws_period[f'C{idx}'] = row[2]
        ws_period[f'D{idx}'] = row[3]
        ws_period[f'E{idx}'] = row[4]

    ws_checkins = wb.create_sheet("Check-in Details")
    detail_headers = ['Date', 'Service Type', 'Member', 'Parish', 'Status', 'Time', 'Checked In By']
    for col, title in zip(['A', 'B', 'C', 'D', 'E', 'F', 'G'], detail_headers):
        ws_checkins[f'{col}1'] = title
        ws_checkins[f'{col}1'].fill = header_fill
        ws_checkins[f'{col}1'].font = header_font

    checkins = attendance_qs.order_by('-checked_in_at')
    for idx, checkin in enumerate(checkins, start=2):
        ws_checkins[f'A{idx}'] = checkin.session.session_date
        ws_checkins[f'B{idx}'] = checkin.session.get_service_type_display()
        ws_checkins[f'C{idx}'] = checkin.member.get_full_name()
        ws_checkins[f'D{idx}'] = checkin.member.parish.name if checkin.member.parish_id else ''
        ws_checkins[f'E{idx}'] = checkin.get_status_display()
        ws_checkins[f'F{idx}'] = checkin.checked_in_at.strftime('%H:%M')
        ws_checkins[f'G{idx}'] = checkin.checked_in_by.get_username() if checkin.checked_in_by else ''
    
    # Auto-adjust column widths
    for ws in [ws_overview, ws_period, ws_checkins]:
        for column_cells in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column_cells:
                # Skip merged cells
                if hasattr(cell, 'column_letter'):
                    column_letter = cell.column_letter
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
    
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=attendance_report_{period}.xlsx'
    
    wb.save(response)
    return response


def load_parishes(request):
    """AJAX view to load parishes based on selected diocese"""
    diocese_id = request.GET.get('diocese_id')
    parishes = Parish.objects.filter(diocese_id=diocese_id).order_by('name')
    return JsonResponse(list(parishes.values('id', 'name')), safe=False)


class MemberListView(LoginRequiredMixin, ListView):
    model = Member
    template_name = 'church/members.html'
    context_object_name = 'members'
    paginate_by = 25
    
    def get_queryset(self):
        queryset = Member.objects.select_related('parish', 'parish__diocese').all()
        
        
        search = self.request.GET.get('search', '')
        if search:
            queryset = queryset.filter(
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(phone__icontains=search)
            )
        
        return queryset.order_by('last_name', 'first_name')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search'] = self.request.GET.get('search', '')
        return context


class MemberCreateView(LoginRequiredMixin, CreateView):
    model = Member
    form_class = MemberForm
    template_name = 'church/add_member.html'
    success_url = reverse_lazy('church:members')
    
    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, f'{self.object.first_name} {self.object.last_name} has been added successfully.')
        return response
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['parishes'] = Parish.objects.all()
        return context


class MemberUpdateView(LoginRequiredMixin, UpdateView):
    model = Member
    form_class = MemberForm
    template_name = 'church/edit_member.html'
    success_url = reverse_lazy('church:members')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['parishes'] = Parish.objects.all()
        return context


class RoleListView(LoginRequiredMixin, ListView):
    model = Role
    template_name = 'church/roles.html'
    context_object_name = 'roles'
    paginate_by = 20


class RoleCreateView(LoginRequiredMixin, CreateView):
    model = Role
    form_class = RoleForm
    template_name = 'church/add_role.html'
    success_url = reverse_lazy('church:roles')


@login_required
def assign_role(request, member_id):
    member = get_object_or_404(Member, id=member_id)
    
    if request.method == 'POST':
        form = MemberRoleForm(request.POST)
        if form.is_valid():
            member_role = form.save(commit=False)
            member_role.member = member
            member_role.save()
            return redirect('church:member_roles', member_id=member_id)
    else:
        form = MemberRoleForm()
    
    roles = Role.objects.all()
    return render(request, 'church/assign_role.html', {
        'member': member,
        'member_id': member_id,
        'roles': roles,
        'form': form
    })


@login_required
def member_roles(request, member_id):
    member = get_object_or_404(Member, id=member_id)
    roles = MemberRole.objects.filter(member=member).select_related('role')
    today = timezone.localdate()
    
    return render(request, 'church/member_roles.html', {
        'member': member,
        'member_id': member_id,
        'roles': roles,
        'today': today,
    })


@login_required
def member_profile(request, member_id):
    """Display member profile with sacramental records"""
    member = get_object_or_404(Member, id=member_id)
   
    try:
        baptism = member.baptism
    except Baptism.DoesNotExist:
        baptism = None
    
    try:
        confirmation = member.confirmation
    except Confirmation.DoesNotExist:
        confirmation = None
   
    marriages = Marriage.objects.filter(
        Q(bride=member) | Q(groom=member)
    )
    
    context = {
        'member': member,
        'baptism': baptism,
        'confirmation': confirmation,
        'marriages': marriages,
    }
    
    return render(request, 'church/member_profile.html', context)


@login_required
def baptism_list(request):
    """List all baptism records with search and pagination"""
    baptisms = Baptism.objects.select_related('member__parish__diocese').all()
    
    # Search
    search = request.GET.get('search', '')
    if search:
        baptisms = baptisms.filter(
            Q(member__first_name__icontains=search) |
            Q(member__last_name__icontains=search) |
            Q(certificate_number__icontains=search)
        )
    
  
    from_date = request.GET.get('from_date', '')
    to_date = request.GET.get('to_date', '')
    if from_date:
        baptisms = baptisms.filter(baptism_date__gte=from_date)
    if to_date:
        baptisms = baptisms.filter(baptism_date__lte=to_date)
    
    baptisms = baptisms.order_by('-baptism_date')
    
    from django.core.paginator import Paginator
    paginator = Paginator(baptisms, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'church/baptism_list.html', {
        'page_obj': page_obj,
        'search': search,
        'from_date': from_date,
        'to_date': to_date,
    })


@login_required
def add_baptism(request):
    """Create a new baptism record from frontend."""
    if request.method == 'POST':
        form = BaptismForm(request.POST)
        if form.is_valid():
            baptism = form.save()
            messages.success(request, f'Baptism for {baptism.member.get_full_name()} added successfully.')
            return redirect('church:baptism_list')
    else:
        form = BaptismForm()

    return render(request, 'church/add_baptism.html', {'form': form})


@login_required
def confirmation_list(request):
    """List all confirmation records with search and pagination"""
    confirmations = Confirmation.objects.select_related('member__parish__diocese').all()
    
    # Search
    search = request.GET.get('search', '')
    if search:
        confirmations = confirmations.filter(
            Q(member__first_name__icontains=search) |
            Q(member__last_name__icontains=search) |
            Q(certificate_number__icontains=search)
        )
    
   
    from_date = request.GET.get('from_date', '')
    to_date = request.GET.get('to_date', '')
    if from_date:
        confirmations = confirmations.filter(confirmation_date__gte=from_date)
    if to_date:
        confirmations = confirmations.filter(confirmation_date__lte=to_date)
    
    confirmations = confirmations.order_by('-confirmation_date')
    
  
    from django.core.paginator import Paginator
    paginator = Paginator(confirmations, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'church/confirmation_list.html', {
        'page_obj': page_obj,
        'search': search,
        'from_date': from_date,
        'to_date': to_date,
    })


@login_required
def add_confirmation(request):
    """Create a new confirmation record from frontend."""
    if request.method == 'POST':
        form = ConfirmationForm(request.POST)
        if form.is_valid():
            confirmation = form.save()
            messages.success(request, f'Confirmation for {confirmation.member.get_full_name()} added successfully.')
            return redirect('church:confirmation_list')
    else:
        form = ConfirmationForm()

    return render(request, 'church/add_confirmation.html', {'form': form})


@login_required
def marriage_list(request):
    """List all marriage records with search and pagination"""
    marriages = Marriage.objects.select_related(
        'bride__parish__diocese', 
        'groom__parish__diocese'
    ).all()
    
 
    search = request.GET.get('search', '')
    if search:
        marriages = marriages.filter(
            Q(bride__first_name__icontains=search) |
            Q(bride__last_name__icontains=search) |
            Q(groom__first_name__icontains=search) |
            Q(groom__last_name__icontains=search) |
            Q(certificate_number__icontains=search)
        )
    

    from_date = request.GET.get('from_date', '')
    to_date = request.GET.get('to_date', '')
    if from_date:
        marriages = marriages.filter(marriage_date__gte=from_date)
    if to_date:
        marriages = marriages.filter(marriage_date__lte=to_date)
    
    marriages = marriages.order_by('-marriage_date')
    
 
    from django.core.paginator import Paginator
    paginator = Paginator(marriages, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'church/marriage_list.html', {
        'page_obj': page_obj,
        'search': search,
        'from_date': from_date,
        'to_date': to_date,
    })


@login_required
def add_marriage(request):
    """Create a new marriage record from frontend."""
    if request.method == 'POST':
        form = MarriageForm(request.POST)
        if form.is_valid():
            marriage = form.save()
            messages.success(
                request,
                f'Marriage record for {marriage.groom.get_full_name()} and {marriage.bride.get_full_name()} added successfully.',
            )
            return redirect('church:marriage_list')
    else:
        form = MarriageForm()

    return render(request, 'church/add_marriage.html', {'form': form})


def generate_qr_code(data):
    """Generate QR code and return as bytes"""
    qr = qrcode.QRCode(version=1, box_size=10, border=2)
    qr.add_data(data)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    
    buffer = BytesIO()
    img.save(buffer, format='PNG')
    buffer.seek(0)
    return buffer


@login_required
def generate_baptism_certificate(request, baptism_id):
    """Generate baptism certificate PDF with QR code"""
    baptism = get_object_or_404(Baptism, id=baptism_id)
    member = baptism.member
    
   
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Baptism_Certificate_{baptism.certificate_number}.pdf"'
    
 
    p = canvas.Canvas(response, pagesize=A4)
    width, height = A4
    
    
    p.setStrokeColorRGB(0.2, 0.4, 0.6)
    p.setLineWidth(3)
    p.rect(40, 40, width - 80, height - 80, stroke=1, fill=0)
    p.setLineWidth(1)
    p.rect(45, 45, width - 90, height - 90, stroke=1, fill=0)
    

    from django.conf import settings
    import os
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'images', 'EAR LOGO.png')
    if os.path.exists(logo_path):
        p.drawImage(logo_path, width/2 - 30, height - 120, width=60, height=60, preserveAspectRatio=True, mask='auto')
    
   
    p.setFont("Helvetica-Bold", 22)
    p.drawCentredString(width/2, height - 140, "ANGLICAN CHURCH OF RWANDA")
    
    p.setFont("Helvetica-Bold", 18)
    p.drawCentredString(width/2, height - 165, "CERTIFICATE OF BAPTISM")
    
   
    p.setStrokeColorRGB(0.2, 0.4, 0.6)
    p.setLineWidth(2)
    p.line(100, height - 175, width - 100, height - 175)
    
    
    p.setFont("Helvetica", 9)
    p.setFillColorRGB(0.3, 0.3, 0.3)
    p.drawCentredString(width/2, height - 190, f"Certificate No: {baptism.certificate_number}")
    
    
    p.setFillColorRGB(0, 0, 0)
    y = height - 230
    p.setFont("Helvetica", 12)
    p.drawCentredString(width/2, y, "This is to certify that")
    y -= 35
    
    p.setFont("Helvetica-Bold", 18)
    p.drawCentredString(width/2, y, member.get_full_name().upper())
    y -= 30
    
    p.setFont("Helvetica-Oblique", 10)
    p.setFillColorRGB(0.3, 0.3, 0.3)
    birth_info = f"Born: {member.date_of_birth.strftime('%B %d, %Y') if member.date_of_birth else 'N/A'} | Gender: {member.get_gender_display()}"
    p.drawCentredString(width/2, y, birth_info)
    y -= 35
    
    p.setFillColorRGB(0, 0, 0)
    p.setFont("Helvetica", 12)
    p.drawCentredString(width/2, y, "was received into the Church of God through the Sacrament of Baptism")
    y -= 30
    
    p.setFont("Helvetica-Bold", 14)
    p.drawCentredString(width/2, y, baptism.baptism_date.strftime('%B %d, %Y'))
    y -= 35
    
    p.setFont("Helvetica", 11)
    p.drawCentredString(width/2, y, f"at {baptism.church_name if baptism.church_name else baptism.parish.name}")
    y -= 20
    p.drawCentredString(width/2, y, f"Parish of {baptism.parish.name}, Diocese of {member.parish.diocese.name}")
    y -= 30
    
    p.setFont("Helvetica-Oblique", 11)
    p.drawCentredString(width/2, y, f"Officiating Priest: {baptism.officiating_priest}")
    y -= 45
    
   
    p.setFillColorRGB(0, 0, 0)
    p.setFont("Helvetica-Bold", 11)
    p.drawString(90, y, "Godparents:")
    y -= 25
    
   
    y -= 10
    box_top = y + 10
    
    p.setFont("Helvetica", 10)
    p.drawString(100, y, f"• {baptism.godparent1_name} ({baptism.get_godparent1_gender_display()})")
    y -= 15
    p.drawString(100, y, f"• {baptism.godparent2_name} ({baptism.get_godparent2_gender_display()})")
    y -= 15
    p.drawString(100, y, f"• {baptism.godparent3_name} ({baptism.get_godparent3_gender_display()})")
    
 
    y -= 10
    box_height = box_top - y
    p.setStrokeColorRGB(0.7, 0.7, 0.7)
    p.setLineWidth(0.5)
    p.rect(80, y, width - 160, box_height, stroke=1, fill=0)
    y -= 30
    
    # Signature lines
    sig_y = 180
    p.setFont("Helvetica", 10)
    
  
    p.line(80, sig_y, 220, sig_y)
    p.drawCentredString(150, sig_y - 15, "Officiating Priest")
    p.setFont("Helvetica", 8)
    p.drawCentredString(150, sig_y - 28, "Signature & Date")
    
   
    p.setFont("Helvetica", 10)
    p.line(width - 220, sig_y, width - 80, sig_y)
    p.drawCentredString(width - 150, sig_y - 15, "Parish Priest/Registrar")
    p.setFont("Helvetica", 8)
    p.drawCentredString(width - 150, sig_y - 28, "Signature & Date")
    
 
    p.setStrokeColorRGB(0.7, 0.7, 0.7)
    p.setLineWidth(0.5)
    p.setDash(2, 2)
    p.circle(width/2, sig_y - 15, 30, stroke=1, fill=0)
    p.setDash()
    p.setFont("Helvetica", 7)
    p.setFillColorRGB(0.5, 0.5, 0.5)
    p.drawCentredString(width/2, sig_y - 40, "OFFICIAL SEAL")
    
 
    verification_url = request.build_absolute_uri(f'/verify-certificate/?cert={baptism.certificate_number}&type=baptism')
    qr_buffer = generate_qr_code(verification_url)
    qr_image = ImageReader(qr_buffer)
    p.drawImage(qr_image, width - 130, 55, width=70, height=70)
    
    p.setFillColorRGB(0, 0, 0)
    p.setFont("Helvetica", 7)
    p.drawCentredString(width - 95, 45, "Scan to verify")
    

    p.setFont("Helvetica-Oblique", 9)
    p.setFillColorRGB(0.2, 0.4, 0.6)
    p.drawCentredString(width/2, 75, "Given under the seal of the Anglican Church of Rwanda")
    p.setFont("Helvetica", 8)
    p.setFillColorRGB(0.3, 0.3, 0.3)
    from datetime import date
    p.drawCentredString(width/2, 60, f"Issued on: {date.today().strftime('%B %d, %Y')}")
    
    p.showPage()
    p.save()
    
    return response


@login_required
def generate_confirmation_certificate(request, confirmation_id):
    """Generate confirmation certificate PDF with QR code"""
    confirmation = get_object_or_404(Confirmation, id=confirmation_id)
    member = confirmation.member
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Confirmation_Certificate_{confirmation.certificate_number}.pdf"'
    
    p = canvas.Canvas(response, pagesize=A4)
    width, height = A4
    
  
    p.setStrokeColorRGB(0.2, 0.4, 0.6)
    p.setLineWidth(3)
    p.rect(40, 40, width - 80, height - 80, stroke=1, fill=0)
    p.setLineWidth(1)
    p.rect(45, 45, width - 90, height - 90, stroke=1, fill=0)
    
    
    from django.conf import settings
    import os
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'images', 'EAR LOGO.png')
    if os.path.exists(logo_path):
        p.drawImage(logo_path, width/2 - 30, height - 120, width=60, height=60, preserveAspectRatio=True, mask='auto')
    
   
    p.setFont("Helvetica-Bold", 22)
    p.drawCentredString(width/2, height - 140, "ANGLICAN CHURCH OF RWANDA")
    
    p.setFont("Helvetica-Bold", 18)
    p.drawCentredString(width/2, height - 165, "CERTIFICATE OF CONFIRMATION")
    
    
    
    p.setStrokeColorRGB(0.2, 0.4, 0.6)
    p.setLineWidth(2)
    p.line(100, height - 175, width - 100, height - 175)
    
    p.setFont("Helvetica", 9)
    p.setFillColorRGB(0.3, 0.3, 0.3)
    p.drawCentredString(width/2, height - 190, f"Certificate No: {confirmation.certificate_number}")
    
   
   
    p.setFillColorRGB(0, 0, 0)
    y = height - 230
    p.setFont("Helvetica", 12)
    p.drawCentredString(width/2, y, "This is to certify that")
    y -= 35
    
    p.setFont("Helvetica-Bold", 18)
    p.drawCentredString(width/2, y, member.get_full_name().upper())
    y -= 30
    
    p.setFont("Helvetica-Oblique", 10)
    p.setFillColorRGB(0.3, 0.3, 0.3)
    birth_info = f"Born: {member.date_of_birth.strftime('%B %d, %Y') if member.date_of_birth else 'N/A'} | Gender: {member.get_gender_display()}"
    p.drawCentredString(width/2, y, birth_info)
    y -= 35
    
    p.setFillColorRGB(0, 0, 0)
    p.setFont("Helvetica", 12)
    p.drawCentredString(width/2, y, "was confirmed in the Christian faith and received the laying on of hands")
    y -= 30
    
    p.setFont("Helvetica-Bold", 14)
    p.drawCentredString(width/2, y, confirmation.confirmation_date.strftime('%B %d, %Y'))
    y -= 35
    
    p.setFont("Helvetica", 11)
    p.drawCentredString(width/2, y, f"at {confirmation.church_name}")
    y -= 20
    p.drawCentredString(width/2, y, f"Parish of {member.parish.name}, Diocese of {member.parish.diocese.name}")
    y -= 30
    
    p.setFont("Helvetica-Oblique", 11)
    p.drawCentredString(width/2, y, f"Right Reverend: {confirmation.bishop_name}")
    y -= 45
    
    
    p.setFillColorRGB(0, 0, 0)
    p.setFont("Helvetica-Bold", 11)
    p.drawString(90, y, "Sponsors:")
    y -= 25
    
    y -= 10
    box_top = y + 10
    
    p.setFont("Helvetica", 10)
    p.drawString(100, y, f"• {confirmation.sponsor1_name} ({confirmation.get_sponsor1_gender_display()})")
    y -= 15
    p.drawString(100, y, f"• {confirmation.sponsor2_name} ({confirmation.get_sponsor2_gender_display()})")
    y -= 15
    p.drawString(100, y, f"• {confirmation.sponsor3_name} ({confirmation.get_sponsor3_gender_display()})")
    if confirmation.sponsor4_name:
        y -= 15
        p.drawString(100, y, f"• {confirmation.sponsor4_name} ({confirmation.get_sponsor4_gender_display()})")
    
    y -= 10
    box_height = box_top - y
    p.setStrokeColorRGB(0.7, 0.7, 0.7)
    p.setLineWidth(0.5)
    p.rect(80, y, width - 160, box_height, stroke=1, fill=0)
    y -= 30
    

    sig_y = 180
    p.setFont("Helvetica", 10)
    
    
    p.line(80, sig_y, 220, sig_y)
    p.drawCentredString(150, sig_y - 15, "Bishop")
    p.setFont("Helvetica", 8)
    p.drawCentredString(150, sig_y - 28, "Signature & Date")
    
 
    p.setFont("Helvetica", 10)
    p.line(width - 220, sig_y, width - 80, sig_y)
    p.drawCentredString(width - 150, sig_y - 15, "Parish Priest/Registrar")
    p.setFont("Helvetica", 8)
    p.drawCentredString(width - 150, sig_y - 28, "Signature & Date")
    

    p.setStrokeColorRGB(0.7, 0.7, 0.7)
    p.setLineWidth(0.5)
    p.setDash(2, 2)
    p.circle(width/2, sig_y - 15, 30, stroke=1, fill=0)
    p.setDash()
    p.setFont("Helvetica", 7)
    p.setFillColorRGB(0.5, 0.5, 0.5)
    p.drawCentredString(width/2, sig_y - 40, "OFFICIAL SEAL")
    
    verification_url = request.build_absolute_uri(f'/verify-certificate/?cert={confirmation.certificate_number}&type=confirmation')
    qr_buffer = generate_qr_code(verification_url)
    qr_image = ImageReader(qr_buffer)
    p.drawImage(qr_image, width - 130, 55, width=70, height=70)
    
    p.setFillColorRGB(0, 0, 0)
    p.setFont("Helvetica", 7)
    p.drawCentredString(width - 95, 45, "Scan to verify")
    
  
    p.setFont("Helvetica-Oblique", 9)
    p.setFillColorRGB(0.2, 0.4, 0.6)
    p.drawCentredString(width/2, 75, "Given under the seal of the Anglican Church of Rwanda")
    p.setFont("Helvetica", 8)
    p.setFillColorRGB(0.3, 0.3, 0.3)
    from datetime import date
    p.drawCentredString(width/2, 60, f"Issued on: {date.today().strftime('%B %d, %Y')}")
    
    p.showPage()
    p.save()
    
    return response


@login_required
def generate_marriage_certificate(request, marriage_id):
    """Generate marriage certificate PDF with QR code"""
    marriage = get_object_or_404(Marriage, id=marriage_id)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Marriage_Certificate_{marriage.certificate_number}.pdf"'
    
    p = canvas.Canvas(response, pagesize=A4)
    width, height = A4
    
   
    p.setStrokeColorRGB(0.2, 0.4, 0.6)
    p.setLineWidth(3)
    p.rect(40, 40, width - 80, height - 80, stroke=1, fill=0)
    p.setLineWidth(1)
    p.rect(45, 45, width - 90, height - 90, stroke=1, fill=0)
    
   
    from django.conf import settings
    import os
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'images', 'EAR LOGO.png')
    if os.path.exists(logo_path):
        p.drawImage(logo_path, width/2 - 30, height - 120, width=60, height=60, preserveAspectRatio=True, mask='auto')
    
  
    p.setFont("Helvetica-Bold", 22)
    p.drawCentredString(width/2, height - 140, "ANGLICAN CHURCH OF RWANDA")
    
    p.setFont("Helvetica-Bold", 18)
    p.drawCentredString(width/2, height - 165, "CERTIFICATE OF MARRIAGE")
    
    
    p.setStrokeColorRGB(0.2, 0.4, 0.6)
    p.setLineWidth(2)
    p.line(100, height - 175, width - 100, height - 175)
    
    p.setFont("Helvetica", 9)
    p.setFillColorRGB(0.3, 0.3, 0.3)
    p.drawCentredString(width/2, height - 190, f"Certificate No: {marriage.certificate_number}")
    
   
    p.setFillColorRGB(0, 0, 0)
    y = height - 230
    p.setFont("Helvetica", 12)
    p.drawCentredString(width/2, y, "This is to certify that")
    y -= 40
    
    
    p.setFont("Helvetica-Bold", 16)
    p.drawCentredString(width/2, y, marriage.bride.get_full_name().upper())
    y -= 20
    p.setFont("Helvetica-Oblique", 10)
    p.setFillColorRGB(0.3, 0.3, 0.3)
    p.drawCentredString(width/2, y, f"Parish: {marriage.bride.parish.name}")
    y -= 35
    
    p.setFillColorRGB(0, 0, 0)
    p.setFont("Helvetica", 12)
    p.drawCentredString(width/2, y, "and")
    y -= 35
    
   
    p.setFont("Helvetica-Bold", 16)
    p.drawCentredString(width/2, y, marriage.groom.get_full_name().upper())
    y -= 20
    p.setFont("Helvetica-Oblique", 10)
    p.setFillColorRGB(0.3, 0.3, 0.3)
    p.drawCentredString(width/2, y, f"Parish: {marriage.groom.parish.name}")
    y -= 35
    
    p.setFillColorRGB(0, 0, 0)
    p.setFont("Helvetica", 12)
    p.drawCentredString(width/2, y, "were joined together in the Sacrament of Holy Matrimony")
    y -= 30
    
    p.setFont("Helvetica-Bold", 14)
    p.drawCentredString(width/2, y, marriage.marriage_date.strftime('%B %d, %Y'))
    y -= 35
    
    p.setFont("Helvetica", 11)
    p.drawCentredString(width/2, y, f"at {marriage.church_name}")
    y -= 30
    
    p.setFont("Helvetica-Oblique", 11)
    p.drawCentredString(width/2, y, f"Officiating Minister: {marriage.minister_name}")
    y -= 45
    
    
    p.setFillColorRGB(0, 0, 0)
    p.setFont("Helvetica-Bold", 11)
    p.drawString(90, y, "Witnesses:")
    y -= 25
    
  
    y -= 10
    box_top = y + 10
    
    p.setFont("Helvetica", 10)
    p.drawString(100, y, f"• {marriage.witness1_name}")
    y -= 15
    p.drawString(100, y, f"• {marriage.witness2_name}")
    
  
    y -= 10
    box_height = box_top - y
    p.setStrokeColorRGB(0.7, 0.7, 0.7)
    p.setLineWidth(0.5)
    p.rect(80, y, width - 160, box_height, stroke=1, fill=0)
    y -= 30
    
   
    sig_y = 210
    p.setFont("Helvetica", 10)
    
   
    p.line(70, sig_y, 180, sig_y)
    p.drawCentredString(125, sig_y - 15, "Officiating Minister")
    p.setFont("Helvetica", 8)
    p.drawCentredString(125, sig_y - 28, "Signature & Date")
    
  
    p.setFont("Helvetica", 10)
    p.line(width/2 - 55, sig_y, width/2 + 55, sig_y)
    p.drawCentredString(width/2, sig_y - 15, "Bride")
    p.setFont("Helvetica", 8)
    p.drawCentredString(width/2, sig_y - 28, "Signature")
    
    
    p.setFont("Helvetica", 10)
    p.line(width - 180, sig_y, width - 70, sig_y)
    p.drawCentredString(width - 125, sig_y - 15, "Groom")
    p.setFont("Helvetica", 8)
    p.drawCentredString(width - 125, sig_y - 28, "Signature")
    

    p.setStrokeColorRGB(0.7, 0.7, 0.7)
    p.setLineWidth(0.5)
    p.setDash(2, 2)
    p.circle(width/2, 150, 30, stroke=1, fill=0)
    p.setDash()
    p.setFont("Helvetica", 7)
    p.setFillColorRGB(0.5, 0.5, 0.5)
    p.drawCentredString(width/2, 115, "OFFICIAL SEAL")
    
  
    verification_url = request.build_absolute_uri(f'/verify-certificate/?cert={marriage.certificate_number}&type=marriage')
    qr_buffer = generate_qr_code(verification_url)
    qr_image = ImageReader(qr_buffer)
    p.drawImage(qr_image, width - 130, 55, width=70, height=70)
    
    p.setFillColorRGB(0, 0, 0)
    p.setFont("Helvetica", 7)
    p.drawCentredString(width - 95, 45, "Scan to verify")
    
    # Footer
    p.setFont("Helvetica-Oblique", 9)
    p.setFillColorRGB(0.2, 0.4, 0.6)
    p.drawCentredString(width/2, 75, "Given under the seal of the Anglican Church of Rwanda")
    p.setFont("Helvetica", 8)
    p.setFillColorRGB(0.3, 0.3, 0.3)
    from datetime import date
    p.drawCentredString(width/2, 60, f"Issued on: {date.today().strftime('%B %d, %Y')}")
    
    p.showPage()
    p.save()
    
    return response


def verify_certificate(request):
    """Public page to verify certificates"""
    cert_number = request.GET.get('cert', '')
    cert_type = request.GET.get('type', '')
    
    result = None
    
    if cert_number:
        if cert_type == 'baptism':
            try:
                baptism = Baptism.objects.select_related('member__parish__diocese').get(certificate_number=cert_number)
                result = {
                    'type': 'Baptism',
                    'valid': True,
                    'member': baptism.member.get_full_name(),
                    'date': baptism.baptism_date,
                    'church': baptism.church_name if baptism.church_name else baptism.parish.name,
                    'minister': baptism.officiating_priest,
                    'parish': baptism.parish.name,
                    'diocese': baptism.parish.diocese.name,
                }
            except Baptism.DoesNotExist:
                result = {'valid': False}
                
        elif cert_type == 'confirmation':
            try:
                confirmation = Confirmation.objects.select_related('member__parish__diocese').get(certificate_number=cert_number)
                result = {
                    'type': 'Confirmation',
                    'valid': True,
                    'member': confirmation.member.get_full_name(),
                    'date': confirmation.confirmation_date,
                    'church': confirmation.church_name if confirmation.church_name else confirmation.parish.name,
                    'bishop': confirmation.confirming_bishop,
                    'parish': confirmation.parish.name,
                    'diocese': confirmation.parish.diocese.name,
                }
            except Confirmation.DoesNotExist:
                result = {'valid': False}
                
        elif cert_type == 'marriage':
            try:
                marriage = Marriage.objects.select_related('bride__parish__diocese', 'groom__parish__diocese').get(certificate_number=cert_number)
                result = {
                    'type': 'Marriage',
                    'valid': True,
                    'bride': marriage.bride.get_full_name(),
                    'groom': marriage.groom.get_full_name(),
                    'date': marriage.marriage_date,
                    'church': marriage.church_name if marriage.church_name else marriage.parish.name,
                    'minister': marriage.officiating_priest,
                    'parish': marriage.parish.name,
                }
            except Marriage.DoesNotExist:
                result = {'valid': False}
    
    return render(request, 'church/verify_certificate.html', {
        'cert_number': cert_number,
        'cert_type': cert_type,
        'result': result
    })


@login_required
def attendance_sessions(request):
    if hasattr(request.user, 'member'):
        return redirect('church:member_portal')

    if request.method == 'POST':
        form = ServiceSessionForm(request.POST)
        if form.is_valid():
            session = form.save(commit=False)
            session.created_by = request.user
            session.save()
            messages.success(request, 'Attendance session created successfully.')
            return redirect('church:attendance_desk', session_id=session.id)
    else:
        form = ServiceSessionForm(initial={'session_date': timezone.localdate()})

    sessions = ServiceSession.objects.select_related('created_by').prefetch_related('attendance_records')[:20]

    session_rows = []
    for session in sessions:
        present_count = session.attendance_records.filter(status='PRESENT').count()
        late_count = session.attendance_records.filter(status='LATE').count()
        total_count = present_count + late_count
        session_rows.append({
            'session': session,
            'present_count': present_count,
            'late_count': late_count,
            'total_count': total_count,
        })

    return render(request, 'church/attendance_sessions.html', {
        'form': form,
        'session_rows': session_rows,
    })


@login_required
def attendance_desk(request, session_id):
    if hasattr(request.user, 'member'):
        return redirect('church:member_portal')

    session = get_object_or_404(ServiceSession, id=session_id)

    if request.method == 'POST':
        member_id = request.POST.get('member_id')
        member = get_object_or_404(Member, id=member_id)
        form = AttendanceCheckInForm(request.POST)

        if form.is_valid():
            status = form.cleaned_data['status']
            record, created = AttendanceRecord.objects.get_or_create(
                session=session,
                member=member,
                defaults={
                    'status': status,
                    'checked_in_by': request.user,
                },
            )

            if created:
                messages.success(request, f'{member.get_full_name()} checked in as {record.get_status_display()}.')
            else:
                record.status = status
                record.checked_in_by = request.user
                record.save(update_fields=['status', 'checked_in_by'])
                messages.info(request, f'{member.get_full_name()} attendance was updated to {record.get_status_display()}.')

            return redirect(f"{reverse('church:attendance_desk', kwargs={'session_id': session.id})}?q={request.GET.get('q', '')}")
    else:
        form = AttendanceCheckInForm(initial={'status': 'PRESENT'})

    query = request.GET.get('q', '').strip()
    members = Member.objects.select_related('parish__diocese').all().order_by('last_name', 'first_name')
    if query:
        members = members.filter(
            Q(first_name__icontains=query)
            | Q(last_name__icontains=query)
            | Q(phone__icontains=query)
            | Q(id__icontains=query)
        )
    members = members[:30]

    checked_map = {
        record.member_id: record
        for record in AttendanceRecord.objects.filter(session=session).select_related('member', 'checked_in_by')
    }

    members_with_status = []
    for member in members:
        members_with_status.append({
            'member': member,
            'record': checked_map.get(member.id),
        })

    checkins = AttendanceRecord.objects.filter(session=session).select_related('member', 'checked_in_by').order_by('-checked_in_at')[:20]

    present_count = AttendanceRecord.objects.filter(session=session, status='PRESENT').count()
    late_count = AttendanceRecord.objects.filter(session=session, status='LATE').count()

    return render(request, 'church/attendance_desk.html', {
        'session': session,
        'form': form,
        'query': query,
        'members_with_status': members_with_status,
        'checkins': checkins,
        'present_count': present_count,
        'late_count': late_count,
        'total_checked_in': present_count + late_count,
    })


@login_required
def export_members_excel(request):
    """Export members list to Excel"""
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Members"
    
    # Add headers
    headers = ['First Name', 'Last Name', 'Gender', 'Date of Birth', 'Phone', 'Diocese', 'Parish', 'Member Since']
    ws.append(headers)
    
    # Style headers
    for cell in ws[1]:
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Get members
    members = Member.objects.select_related('parish', 'parish__diocese').order_by('last_name', 'first_name')
    
    # Add data
    for member in members:
        ws.append([
            member.first_name,
            member.last_name,
            member.get_gender_display() if member.gender else '-',
            member.date_of_birth.strftime('%Y-%m-%d') if member.date_of_birth else '-',
            member.phone if member.phone else '-',
            member.parish.diocese.name,
            member.parish.name,
            member.created_at.strftime('%Y-%m-%d')
        ])
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=members_list.xlsx'
    wb.save(response)
    
    return response


@login_required
def export_members_pdf(request):
    """Export members list to PDF"""
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=members_list.pdf'
    
    # Create PDF
    p = canvas.Canvas(response, pagesize=A4)
    width, height = A4
    
    # Title
    p.setFont("Helvetica-Bold", 16)
    p.drawString(50, height - 50, "Members List")
    
    # Get members
    members = Member.objects.select_related('parish', 'parish__diocese').order_by('last_name', 'first_name')
    
    # Prepare data for table
    data = [['Name', 'Gender', 'Phone', 'Diocese', 'Parish', 'Member Since']]
    for member in members:
        data.append([
            f"{member.first_name} {member.last_name}",
            member.get_gender_display() if member.gender else '-',
            member.phone if member.phone else '-',
            member.parish.diocese.name,
            member.parish.name,
            member.created_at.strftime('%b %d, %Y')
        ])
    
    # Create table
    table = Table(data, colWidths=[100, 40, 70, 80, 100, 70])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))
    
    # Draw table
    table.wrapOn(p, width, height)
    table.drawOn(p, 50, height - 100 - (len(data) * 20))
    
    # Add footer
    p.setFont("Helvetica", 8)
    p.drawString(50, 30, f"Total Members: {members.count()}")
    from datetime import datetime
    p.drawString(width - 150, 30, f"Generated: {datetime.now().strftime('%Y-%m-%d')}")
    
    p.showPage()
    p.save()
    
    return response
